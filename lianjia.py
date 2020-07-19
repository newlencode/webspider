import re
import requests
from openpyxl import load_workbook,Workbook

def get_data(url,headers):
    response = requests.get(url,headers = headers)
    html = response.text
    #获取楼盘名
    pattern1 = '}">(.*)</a>\n.*<span class=".*" style=".*:'
    names = re.findall(pattern1,html)
    #获取楼盘地址
    pattern2 = '>(.*)</a>\n.*</div>\n.*<a class="resblock-room" href='
    address = re.findall(pattern2,html)
    #获取户型
    pattern3 ='<a class="resblock-room" href=([\s\S]*?) <div class="resblock-area">'
    res = re.findall(pattern3,html)
    typies = []
    pattern3_1 = '<span>(.*?)</span>'
    for i in range(len(res)-1):
        a = re.findall(pattern3_1,res[i])
        b = ','.join(a)
        typies.append(b)
    #获取建筑面积
    pattern4 = '<div class="resblock-area">\n.*<span>(.*?)</span>'
    aeras = re.findall(pattern4,html)
    aeras.pop()
    #获取价格
    pattern5 = '<div class="main-price">\n.*<span class="number">(.*)</span>'
    prices = re.findall(pattern5,html)
    return names,address,typies,aeras,prices

def save_data(res,j):
    wb = load_workbook('lianjia.xlsx')
    ws = wb.active
    for i in range(len(res[0])):
        ws.cell(i + 2 + j*10, 1).value = res[0][i]
        ws.cell(i + 2 + j*10, 2).value = res[1][i]
        ws.cell(i + 2 + j*10, 3).value = str(res[2][i])
        ws.cell(i + 2 + j*10, 4).value = res[3][i]
        ws.cell(i + 2 + j*10, 5).value = res[4][i]
    wb.save('lianjia.xlsx')
def main():
    wb = Workbook()
    ws = wb.active
    xlhead = ['楼盘名','地址','户型','建筑面积','均价(元/平)']
    for i in range(len(xlhead)):
        ws.cell(1, i+1).value = xlhead[i]
    wb.save('lianjia.xlsx')
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.61 Safari/537.36'}
    num = 1
    for j in range(10):
        url = 'https://sz.fang.lianjia.com/loupan/pg' + str(num)
        res = get_data(url, headers)
        save_data(res,j)
        num += 1

if __name__ == "__main__":
    main()

